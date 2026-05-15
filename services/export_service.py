"""
Export service for NFC Campus E-Wallet System.

Provides Excel export functionality for reports.
"""

from sqlalchemy.orm import Session
from typing import Optional
import io
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed. Excel export will not be available.")

from services.report_service import ReportService
from models.transaction import Transaction
from models.booth import Booth
from models.product import Product

logger = logging.getLogger(__name__)


class ExportService:
    """导出服务类"""
    
    def __init__(self, db: Session):
        """
        Initialize export service.
        
        Args:
            db: Database session
        """
        self.db = db
        self.report_service = ReportService(db)
    
    def export_to_excel(
        self,
        report_type: str,
        event_id: Optional[int] = None
    ) -> bytes:
        """
        导出报表为 Excel 文件。
        
        Args:
            report_type: 报表类型（summary/booths/products/transactions）
            event_id: 活动ID（可选）
        
        Returns:
            bytes: Excel 文件内容
        
        Raises:
            ValueError: 如果报表类型无效或 openpyxl 未安装
        """
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl is not installed. Please install it to use Excel export.")
        
        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认工作表
        
        if report_type == "summary":
            self._export_summary_report(wb, event_id)
        elif report_type == "booths":
            self._export_booth_report(wb, event_id)
        elif report_type == "products":
            self._export_product_report(wb, event_id)
        elif report_type == "transactions":
            self._export_transaction_report(wb, event_id)
        else:
            raise ValueError(f"Invalid report type: {report_type}")
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _export_summary_report(self, wb: Workbook, event_id: Optional[int] = None):
        """导出总览统计报表"""
        ws = wb.create_sheet("总览统计")
        
        # 获取数据
        summary = self.report_service.get_summary_report(event_id)
        
        # 设置标题样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # 写入标题
        ws['A1'] = "总览统计报表"
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        
        # 写入数据
        data = [
            ["指标", "数值"],
            ["总发放额度（元）", summary.total_issued],
            ["总充值额（元）", summary.total_recharged],
            ["总消费额（元）", summary.total_consumed],
            ["总退款额（元）", summary.total_refunded],
            ["净消费额（元）", summary.net_consumed],
            ["总交易笔数", summary.total_transactions],
            ["参与者数量", summary.participant_count],
            ["摊位数量", summary.booth_count]
        ]
        
        for row_idx, row_data in enumerate(data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # 表头
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def _export_booth_report(self, wb: Workbook, event_id: Optional[int] = None):
        """导出摊位报表"""
        ws = wb.create_sheet("摊位报表")
        
        # 获取数据
        booth_report = self.report_service.get_booth_report(event_id)
        
        # 设置样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # 写入标题
        ws['A1'] = "摊位报表"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        
        # 写入表头
        headers = [
            "摊位ID", "摊位名称", "班级名称", "营业额（元）", "退款额（元）",
            "净收入（元）", "销量（笔）", "总成本（元）", "利润（元）", "利润率（%）"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_idx, booth in enumerate(booth_report.booths, start=4):
            ws.cell(row=row_idx, column=1, value=booth.booth_id)
            ws.cell(row=row_idx, column=2, value=booth.booth_name)
            ws.cell(row=row_idx, column=3, value=booth.class_name)
            ws.cell(row=row_idx, column=4, value=booth.revenue)
            ws.cell(row=row_idx, column=5, value=booth.refund_amount)
            ws.cell(row=row_idx, column=6, value=booth.net_revenue)
            ws.cell(row=row_idx, column=7, value=booth.sales_count)
            ws.cell(row=row_idx, column=8, value=booth.total_cost)
            ws.cell(row=row_idx, column=9, value=booth.profit)
            ws.cell(row=row_idx, column=10, value=booth.profit_margin)
        
        # 调整列宽
        for col_idx in range(1, 11):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    def _export_product_report(self, wb: Workbook, event_id: Optional[int] = None):
        """导出商品报表"""
        ws = wb.create_sheet("商品报表")
        
        # 获取数据
        product_report = self.report_service.get_product_report(event_id)
        
        # 设置样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # 写入标题
        ws['A1'] = "商品报表"
        ws['A1'].font = title_font
        ws.merge_cells('A1:I1')
        
        # 写入表头
        headers = [
            "商品ID", "商品名称", "摊位ID", "摊位名称", "销量（件）",
            "收入（元）", "总成本（元）", "利润（元）", "利润率（%）"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_idx, product in enumerate(product_report.products, start=4):
            ws.cell(row=row_idx, column=1, value=product.product_id)
            ws.cell(row=row_idx, column=2, value=product.product_name)
            ws.cell(row=row_idx, column=3, value=product.booth_id)
            ws.cell(row=row_idx, column=4, value=product.booth_name)
            ws.cell(row=row_idx, column=5, value=product.sales_quantity)
            ws.cell(row=row_idx, column=6, value=product.revenue)
            ws.cell(row=row_idx, column=7, value=product.total_cost)
            ws.cell(row=row_idx, column=8, value=product.profit)
            ws.cell(row=row_idx, column=9, value=product.profit_margin)
        
        # 调整列宽
        for col_idx in range(1, 10):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    def _export_transaction_report(self, wb: Workbook, event_id: Optional[int] = None):
        """导出交易流水报表"""
        ws = wb.create_sheet("交易流水")
        
        # 查询交易数据
        query = self.db.query(Transaction)
        if event_id:
            query = query.filter(Transaction.event_id == event_id)
        
        transactions = query.order_by(Transaction.created_at.desc()).limit(10000).all()
        
        # 设置样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # 写入标题
        ws['A1'] = "交易流水报表"
        ws['A1'].font = title_font
        ws.merge_cells('A1:L1')
        
        # 写入表头
        headers = [
            "交易ID", "交易类型", "金额（元）", "交易前余额（元）", "交易后余额（元）",
            "参与者卡号", "摊位ID", "商品ID", "操作员ID", "备注", "交易时间", "活动ID"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_idx, txn in enumerate(transactions, start=4):
            ws.cell(row=row_idx, column=1, value=txn.id)
            ws.cell(row=row_idx, column=2, value=txn.type)
            ws.cell(row=row_idx, column=3, value=float(txn.amount))
            ws.cell(row=row_idx, column=4, value=float(txn.balance_before))
            ws.cell(row=row_idx, column=5, value=float(txn.balance_after))
            ws.cell(row=row_idx, column=6, value=txn.card_uid)
            ws.cell(row=row_idx, column=7, value=txn.booth_id)
            ws.cell(row=row_idx, column=8, value=txn.product_id)
            ws.cell(row=row_idx, column=9, value=txn.booth_operator_id)
            ws.cell(row=row_idx, column=10, value=txn.remark)
            ws.cell(row=row_idx, column=11, value=txn.created_at.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_idx, column=12, value=txn.event_id)
        
        # 调整列宽
        for col_idx in range(1, 13):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    def export_refund_adjustments(self, event_id: Optional[int] = None) -> bytes:
        """
        导出退款/更正清单。
        
        Args:
            event_id: 活动ID（可选）
        
        Returns:
            bytes: Excel 文件内容
        """
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl is not installed. Please install it to use Excel export.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "退款更正清单"
        
        # 查询退款和调整记录
        query = self.db.query(Transaction).filter(
            Transaction.type.in_(['refund', 'adjust'])
        )
        if event_id:
            query = query.filter(Transaction.event_id == event_id)
        
        transactions = query.order_by(Transaction.created_at.desc()).all()
        
        # 设置样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        
        # 写入标题
        ws['A1'] = "退款/更正清单"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        
        # 写入表头
        headers = [
            "交易ID", "类型", "金额（元）", "参与者卡号", "关联交易ID",
            "操作员ID", "备注", "交易时间", "活动ID", "摊位ID"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_idx, txn in enumerate(transactions, start=4):
            ws.cell(row=row_idx, column=1, value=txn.id)
            ws.cell(row=row_idx, column=2, value=txn.type)
            ws.cell(row=row_idx, column=3, value=float(txn.amount))
            ws.cell(row=row_idx, column=4, value=txn.card_uid)
            ws.cell(row=row_idx, column=5, value=txn.related_txn_id)
            ws.cell(row=row_idx, column=6, value=txn.booth_operator_id)
            ws.cell(row=row_idx, column=7, value=txn.remark)
            ws.cell(row=row_idx, column=8, value=txn.created_at.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_idx, column=9, value=txn.event_id)
            ws.cell(row=row_idx, column=10, value=txn.booth_id)
        
        # 调整列宽
        for col_idx in range(1, 11):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_class_settlement(self, event_id: int) -> bytes:
        """
        导出班级结算单。
        
        Args:
            event_id: 活动ID
        
        Returns:
            bytes: Excel 文件内容
        """
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl is not installed. Please install it to use Excel export.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "班级结算单"
        
        # 获取摊位报表数据
        booth_report = self.report_service.get_booth_report(event_id)
        
        # 设置样式
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # 写入标题
        ws['A1'] = "班级结算单"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        
        # 写入表头
        headers = [
            "班级名称", "摊位名称", "营业额（元）", "退款额（元）",
            "净收入（元）", "总成本（元）", "利润（元）", "利润率（%）"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_idx, booth in enumerate(booth_report.booths, start=4):
            ws.cell(row=row_idx, column=1, value=booth.class_name)
            ws.cell(row=row_idx, column=2, value=booth.booth_name)
            ws.cell(row=row_idx, column=3, value=booth.revenue)
            ws.cell(row=row_idx, column=4, value=booth.refund_amount)
            ws.cell(row=row_idx, column=5, value=booth.net_revenue)
            ws.cell(row=row_idx, column=6, value=booth.total_cost)
            ws.cell(row=row_idx, column=7, value=booth.profit)
            ws.cell(row=row_idx, column=8, value=booth.profit_margin)
        
        # 调整列宽
        for col_idx in range(1, 9):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_leaderboard(self, event_id: int) -> bytes:
        """
        导出排名表。
        
        Args:
            event_id: 活动ID
        
        Returns:
            bytes: Excel 文件内容
        """
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl is not installed. Please install it to use Excel export.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "排名表"
        
        # 获取摊位报表数据并按净收入排序
        booth_report = self.report_service.get_booth_report(event_id)
        sorted_booths = sorted(
            booth_report.booths,
            key=lambda x: x.net_revenue,
            reverse=True
        )
        
        # 设置样式
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # 写入标题
        ws['A1'] = "摊位排名表（按净收入）"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        
        # 写入表头
        headers = [
            "排名", "班级名称", "摊位名称", "净收入（元）", "销量（笔）", "利润率（%）"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for rank, booth in enumerate(sorted_booths, start=1):
            row_idx = rank + 3
            ws.cell(row=row_idx, column=1, value=rank)
            ws.cell(row=row_idx, column=2, value=booth.class_name)
            ws.cell(row=row_idx, column=3, value=booth.booth_name)
            ws.cell(row=row_idx, column=4, value=booth.net_revenue)
            ws.cell(row=row_idx, column=5, value=booth.sales_count)
            ws.cell(row=row_idx, column=6, value=booth.profit_margin)
        
        # 调整列宽
        for col_idx in range(1, 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

    def export_booth_transactions(
        self,
        booth_id: int,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        has_product: Optional[bool] = None
    ) -> bytes:
        """
        导出指定商铺的账目明细。
        
        Args:
            booth_id: 摊位ID
            start_date: 开始日期（可选，ISO格式）
            end_date: 结束日期（可选，ISO格式）
            has_product: 是否关联商品（可选）
        
        Returns:
            bytes: Excel 文件内容
        """
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl is not installed. Please install it to use Excel export.")
        
        # 获取摊位信息
        booth = self.db.query(Booth).filter(Booth.id == booth_id).first()
        if not booth:
            raise ValueError(f"Booth with id {booth_id} not found")
        
        # 构建查询
        query = self.db.query(Transaction).filter(Transaction.booth_id == booth_id)
        
        # 应用日期过滤
        if start_date:
            from datetime import datetime
            start_datetime = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(Transaction.created_at >= start_datetime)
        
        if end_date:
            from datetime import datetime
            end_datetime = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(Transaction.created_at <= end_datetime)
        
        # 应用商品关联过滤
        if has_product is not None:
            if has_product:
                query = query.filter(Transaction.product_id.isnot(None))
            else:
                query = query.filter(Transaction.product_id.is_(None))
        
        transactions = query.order_by(Transaction.created_at.desc()).all()
        
        # 预加载商品名称
        product_ids = set(txn.product_id for txn in transactions if txn.product_id)
        product_name_map = {}
        if product_ids:
            products = self.db.query(Product).filter(Product.id.in_(product_ids)).all()
            product_name_map = {p.id: p.name for p in products}
        
        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "商铺账目明细"
        
        # 设置样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # 写入标题
        booth_title = f"商铺账目明细 - {booth.name}"
        if hasattr(booth, 'class_name') and booth.class_name:
            booth_title += f" ({booth.class_name})"
        ws['A1'] = booth_title
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        
        # 写入汇总信息
        total_income = sum(
            float(txn.amount) for txn in transactions
            if txn.type in ('pay', 'cash_payment')
        )
        total_refund = sum(
            float(txn.amount) for txn in transactions
            if txn.type == 'refund'
        )
        ws['A2'] = f"总笔数: {len(transactions)}  |  收入合计: ¥{total_income:.2f}  |  退款合计: ¥{total_refund:.2f}  |  净收入: ¥{total_income - total_refund:.2f}"
        ws.merge_cells('A2:J2')
        
        # 写入表头
        headers = [
            "交易ID", "交易类型", "金额（元）", "交易前余额（元）", "交易后余额（元）",
            "卡号", "商品名称", "操作员ID", "备注", "交易时间"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 交易类型映射
        type_map = {
            'pay': '支付',
            'refund': '退款',
            'recharge': '充值',
            'cash_payment': '现金收款',
            'correction': '更正',
            'stock_buy': '股票买入',
            'stock_sell': '股票卖出',
            'loan_issue': '垫资发放',
            'loan_fee': '手续费',
        }
        
        # 写入数据
        for row_idx, txn in enumerate(transactions, start=5):
            ws.cell(row=row_idx, column=1, value=txn.id)
            ws.cell(row=row_idx, column=2, value=type_map.get(txn.type, txn.type))
            ws.cell(row=row_idx, column=3, value=float(txn.amount))
            ws.cell(row=row_idx, column=4, value=float(txn.balance_before))
            ws.cell(row=row_idx, column=5, value=float(txn.balance_after))
            ws.cell(row=row_idx, column=6, value=txn.card_uid)
            ws.cell(row=row_idx, column=7, value=product_name_map.get(txn.product_id, '非商品收款') if txn.product_id is None or txn.product_id not in product_name_map else product_name_map[txn.product_id])
            ws.cell(row=row_idx, column=8, value=txn.operator_id)
            ws.cell(row=row_idx, column=9, value=txn.remark)
            ws.cell(row=row_idx, column=10, value=txn.created_at.strftime("%Y-%m-%d %H:%M:%S") if txn.created_at else '')
        
        # 调整列宽
        col_widths = [10, 12, 14, 16, 16, 14, 18, 10, 24, 20]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
